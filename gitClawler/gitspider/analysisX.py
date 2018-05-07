#!usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@author:binbinzhang
@file: analysis.py
@time: 2018/05/07
@email:binbin_Erices@163.com
@function：进行Github代码提交量的数据清洗
"""

from openpyxl import load_workbook,Workbook


# @berif:修改多字节字符乱码的情况
def gbk2utf(in_data , tag):
    if 1 == tag:
        return in_data.encode('gbk').decode('gbk')
    elif 0 == tag:
        return in_data.encode('gbk','ignore').decode('gbk').encode('utf8').decode("utf8")

# @berif:写入工作表
def saveSheet(file,emploment,name,count,notes):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "worksheet"

    length = len(name)

    i=0
    while i< length:
        sheet["A"+str(i+1)] = emploment[i]
        sheet["B"+str(i+1)] = name[i]
        sheet["C"+str(i+1)] = count[i]
        sheet["D"+str(i+1)] = notes[i]
        i += 1

    # 设置提交次数的筛选
    # sheet.auto_filter.ref = "A1:D{}".format(sheet.max_row)
    # sheet.auto_filter.add_sort_condition("C2:C{}".format(sheet.max_row))
    wb.save("./Excel/{}_any.xlsx".format(file))


# @berif:主分析模块
def anay():
    file = input("  请输入需要分析的文件名(输入q!退出程序)：\n")
    if file == "q!":
        print("欢迎下次继续使用GitHub数据分析...")
        exit(0)
    try:
        wb = load_workbook("./Excel/{}.xlsx".format(file))
    except:
        print("文件名不存在，请输入正确的文件名称...")
        return

    '''
    namedict = {u"基础开发部":["张三","李四","王五"],
                u"平台开发部":["张三","李四","王五"],
                u"应用开发部": ["张三","李四","王五"],
                u"产品部":["张三","李四","王五"],
                u"其他":["张三","李四","王五"],
                }
    '''
    namedict = {
        u"基础开发部": [],
        u"平台开发部": [],
        u"应用开发部": [],
        u"产品部": []
        }
    with open("./anayName.json", "r") as obj:
        namedict[u"基础开发部"] = obj.readline()
        namedict[u"平台开发部"] = obj.readline()
        namedict[u"应用开发部"] = obj.readline()
        namedict[u"产品部"] = obj.readline()
        namedict[u"其他"] = obj.readline()
        obj.close()

    sheet = wb["Sheet"]
    all = 0
    res = {}
    records = ['        {}       \n\n'.format(file)]
    for x in range(2, sheet.max_row + 1):
        if not res.get(sheet[x][3].value):
            res[sheet[x][3].value] = {}
        if not res[sheet[x][3].value].get(sheet[x][1].value):
            res[sheet[x][3].value].update({sheet[x][1].value: 1})
        else:
            res[sheet[x][3].value][sheet[x][1].value] += 1

    emploment = ["部门"]
    name_t = ["姓名"]
    count_t = ["提交次数"]
    notes = ["备注"]

    for pro, names in res.items():
        count = 0

        for name, num in names.items():
            count += num
            if gbk2utf(name,0) in namedict[u'基础开发部']:
                emploment.append(u'基础开发部')
                name_t.append(name)
                count_t.append(str(num))
                notes.append('')

            if gbk2utf(name,0) in namedict[u'平台开发部']:
                emploment.append(u'平台开发部')
                name_t.append(name)
                count_t.append(str(num))
                notes.append('')

            if gbk2utf(name,0) in namedict[u'应用开发部']:
                emploment.append(u'应用开发部')
                name_t.append(name)
                count_t.append(str(num))
                notes.append('')

            if gbk2utf(name,0) in namedict[u'产品部']:
                emploment.append(u'产品部')
                name_t.append(name)
                count_t.append(str(num))
                notes.append('')

            if name in namedict[u'其他']:
                emploment.append(u'其他')
                name_t.append(name)
                count_t.append(str(num))
                notes.append('')
        all += count

    # 导出生成的表格
    try:
        saveSheet(file,emploment, name_t, count_t, notes)
    except:
        print("文件已经生成，无需重新生成...")
        ret = input("请把已经生成文件删除,继续生成最新文件...")
        return

    delrepeatData(file)

    workshow(all,file)

    print("Analysis is ending...\n")


# @berif:筛选重复的数据
def delrepeatData(file):
    emploment = ["部门"]
    name_t = ["姓名"]
    count_t = ["提交次数"]
    notes = ["备注"]
    wb = load_workbook("./Excel/{}_any.xlsx".format(file))
    sheet = wb['worksheet']
    for x in range(2, sheet.max_row+1):
        if sheet[x][1].value in name_t:

            # 在列表里面出现的索引
            index = name_t.index(sheet[x][1].value)

            # 更新重复的信息
            tmp = int(count_t[index]) + int(sheet[x][2].value)
            count_t[index] = str(tmp)

        else:
            emploment.append(str(sheet[x][0].value))
            name_t.append(str(sheet[x][1].value))
            count_t.append(sheet[x][2].value)
            notes.append('')
    saveSheet(file, emploment, name_t, count_t, notes)



# 定义函数
# @berif:打印工作量统计信息到屏幕，并且保存到本地
def workshow(num,file):
    wb = load_workbook("./Excel/{}_any.xlsx".format(file))
    sheet = wb['worksheet']
    records = []
    s1 = ''
    s2 = ''
    s3 = ''
    s4 = ''
    s5 = ''
    for x in range(2, sheet.max_row + 1):
        if sheet[x][0].value == u'基础开发部':
            s1 += '\t'+str(sheet[x][1].value) + u"\t提交代码:\t" + str(sheet[x][2].value) + u"次" + "\n"
        elif sheet[x][0].value == u'平台开发部':
            s2 += '\t'+str(sheet[x][1].value) + u"\t提交代码:\t" + str(sheet[x][2].value) + u"次" + "\n"
        elif sheet[x][0].value == u'应用开发部':
            s3 += '\t'+str(sheet[x][1].value) + u"\t提交代码:\t" + str(sheet[x][2].value) + u"次" + "\n"
        elif sheet[x][0].value == u'产品部':
            s4 += '\t'+str(sheet[x][1].value) + u"\t提交代码:\t" + str(sheet[x][2].value) + u"次" + "\n"
        elif sheet[x][0].value == u'其他':
            s5 += '\t'+str(sheet[x][1].value) + u"\t提交代码:\t" + str(sheet[x][2].value) + u"次" + "\n"

    record = u'基础开发部:\n' + s1 + '\n' + u'平台开发部:\n' + s2 + '\n' + u'应用开发部:\n' + s3 + '\n' + u'产品部:\n' + s4 + '\n' + u'其他：:\n' + s5 + '\n'

    print(record)

    records.append(file+'\n\n')
    records.append(record)


    all = u"代码提交总数：" + str(num) + '\n\n\n'
    records.append(all)
    print("*" * 40 + " " * 4 + str(all))

    # 写入文件
    with open("./gitResults.txt", "a", encoding="utf8") as f:
        f.writelines(records)


if __name__ == "__main__":
    print(" *****************************************************************************")
    print(" *                                                                           *")
    print(" *                                                                           *")
    print(" *                      欢迎使用GitHub数据分析                               *")
    print(" *                                                                           *")
    print(" *                                                                           *")
    print(" *****************************************************************************\n")
    while True:
        anay()



