#!usr/bin/env python  
# -*- coding: UTF-8 -*- 
""" 
@author:binbinzhang
@file: analysis.py 
@time: 2018/04/28 
@email:binbin_Erices@163.com
@function： 
"""

from openpyxl import load_workbook

def anay():
    file = input("  请输入需要分析的文件名(输入q!退出程序)：\n")
    if file == "q!":
        print("欢迎下次继续使用GitHub数据分析...")
        exit(0)
    wb = load_workbook("./Excel/{}.xlsx".format(file))
    sheet = wb["Sheet"]
    # print(sheet.max_row)
    all = 0
    res = {}
    records = ['        {}       \n\n'.format(file)]
    for x in range(2, sheet.max_row + 1):
        if not res.get(sheet[x][3].value):
            res[sheet[x][3].value] = {}
        if not res[sheet[x][3].value].get(sheet[x][1].value):
            res[sheet[x][3].value].update({sheet[x][1].value: 1})
        else:
            res[sheet[x][3].value][sheet[x][1].value] += 1
    for pro, names in res.items():
        count = 0
        s = ""
        for name, num in names.items():
            count += num
            s += name + " " + str(num) + u"次 "
        print(pro + u" 提交了" + str(count) + u"次  （" + s + "）")
        record = pro + u" 提交了" + str(count) + u"次  （" + s + "）"+'\n'
        records.append(record)
        all += count
    all = u"代码提交总数："+str(all)+'\n\n\n'
    records.append(all)
    print("*" * 40+" "*4+ str(all))

    #写入文件
    with open("./gitResults.txt","a") as f:
        f.writelines(records)

    print("Analysis is ending...\n")

if __name__ == "__main__":
    print(" *********************************************************************************")
    print(" *                                                                               *")
    print(" *                                                                               *")
    print(" *                      欢迎使用GitHub数据分析                                   *")
    print(" *                                                                               *")
    print(" *                                                                               *")
    print(" *********************************************************************************\n")
    while True:
        anay()





